from selenium import webdriver
from openpyxl import Workbook
from openpyxl import load_workbook

formurllist=[]
Defect=[]

book = load_workbook('C:\DF-CMT.xlsx')
sheet = book.worksheets[0]
sheet['Q1']="Defect Status Comments"
countrow=sheet.max_row
print(countrow)
for i in range(2,countrow+1):
    defect_temp=(sheet['A'+str(i)].value).strip()
    Defect.append(defect_temp[2:])
print(Defect)
for i in range(len(Defect)):
    formurllist.append('https://xxxxxxx/Defects/DispForm.aspx?ID='+str(Defect[i])+'#SPBookmark_xxxxxxx')
print(formurllist)
driver = webdriver.Ie('IEDriverServer.exe')
try:
    for j in range(500,968):
        
            driver.get(formurllist[j])
            for i in range(10,30):
                a=driver.find_element_by_xpath('//*[@id="part1"]/table[1]/tbody/tr['+str(i)+']/td[1]')
                
                if (a.text)=="Defect Status Comments":
                    a=driver.find_element_by_xpath('//*[@id="part1"]/table[1]/tbody/tr['+str(i)+']/td[2]')
                    print(Defect[j])
                    print (a.text)
                    sheet['Q'+str(j+2)] = a.text
                    break
except e:
    print (e)
book.save("DF-CMT-Update2.xlsx")         

