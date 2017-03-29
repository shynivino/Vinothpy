import xlrd
from tkinter.filedialog import askopenfilename
from tkinter import *
import pyautogui as PG
import openpyxl,re ,math
from openpyxl.styles import PatternFill
from openpyxl.styles.colors import YELLOW
from openpyxl.styles import Border, Side
class test:
    def f1(self):
        Tk().withdraw()
        filename = askopenfilename(title="Choose CDA LOSS PROFILE TEMPLATE")
        workbook = xlrd.open_workbook(filename)
        worksheet = workbook.sheet_by_name('Sheet1')
        num_cols = worksheet.ncols - 1
        num_rows = worksheet.nrows - 1
        print (num_cols,num_rows)
        self.f2(num_cols,num_rows,filename)
    def f2(self,num_cols,num_rows,filename1):
         wb = openpyxl.load_workbook(filename=filename1)
         ws = wb.worksheets[0]
         thin_border = Border(left=Side(style='thin'),
                              right=Side(style='thin'),
                              top=Side(style='thin'),
                              bottom=Side(style='thin'))
         ws.cell(row=1, column=num_cols+2).border = thin_border
         ws.cell(row=1, column=num_cols+3).border = thin_border
         ws.cell(row=1, column=num_cols+4).border = thin_border
         ws.cell(row=1, column=num_cols+2).value = "NBN Calculation"
         ws.cell(row=1, column=num_cols+3).value = "Round off values"
         ws.cell(row=1, column=num_cols+4).value = "BOMBOQ Values"
         ws.cell(row=1, column=num_cols+2).fill = PatternFill(fill_type="solid", start_color="b3b3b3")
         ws.cell(row=1, column=num_cols+3).fill = PatternFill(fill_type="solid", start_color="b3b3b3")
         ws.cell(row=1, column=num_cols+4).fill = PatternFill(fill_type="solid", start_color="b3b3b3")

         for i in range(2,num_cols):
             if (re.search('Total Service Addresses',(ws.cell(row=1, column=i).value))):
                print (ws.cell(row=1, column=i).value)
                break
         for k in range(2,num_cols):
             if ('SIO' == (ws.cell(row=1, column=k).value)):
                print (ws.cell(row=1, column=k).value)
                break
         for j in range(2,num_rows+2):
            if (ws.cell(row=j, column=i).value) > (ws.cell(row=j, column=k).value):
              ws.cell(row=j, column=num_cols+2).value = ((ws.cell(row=j, column=i).value)*1.05)/48
              ws.cell(row=j, column=num_cols + 2).border = thin_border
              ws.cell(row=j, column=num_cols+3).value = math.ceil(((ws.cell(row=j, column=i).value)*1.05)/48)
              ws.cell(row=j, column=num_cols + 3).border = thin_border

            else:
              ws.cell(row=j, column=num_cols+2).value = ((ws.cell(row=j, column=k).value)*1.05)/48
              ws.cell(row=j, column=num_cols + 2).border = thin_border
              ws.cell(row=j, column=num_cols+3).value = math.ceil(((ws.cell(row=j, column=k).value)*1.05)/48)
              ws.cell(row=j, column=num_cols +3).border = thin_border
         self.f3(ws,num_rows,num_cols+4,wb,thin_border,filename1)

    def f3(self,ws,num_rows1,num_cols,wb,thin_border,filename1):
        Tk().withdraw()
        filename=askopenfilename(title="Choose BOMBAQ File")
        workbook = xlrd.open_workbook(filename)
        worksheet = workbook.sheet_by_name('BOM')
        print (filename,worksheet)
        num_rows = worksheet.nrows - 1
        for i in range(num_rows):
            if (re.search("FTTN Line Card",(worksheet.cell(i,1).value))):
                print((worksheet.cell(i,1).value),i+1)
                break

        for x in range(2,num_rows1+2):
            for j in range(7,100):
                temp=(ws.cell(row=x, column=1).value)
                source=worksheet.cell(0,j).value
                source = source[-2:]
                source1 = worksheet.cell(1,j).value
                source1 = source1[-2:]
                source2 = worksheet.cell(2,j).value
                source2 = source2[-2:]
                temp = temp[8:10]
                if(source == temp)  or (source1 == temp) or (temp == source2) :
                   ws.cell(row=x, column=num_cols).value = worksheet.cell(i+1,j).value
                   ws.cell(row=x, column=num_cols).border = thin_border

                   break


        wb.save(filename1)

        for k in range(2, num_cols):
            if ('FNO TYPE' == (ws.cell(row=1, column=k).value)):
                print(ws.cell(row=1, column=k).value)
                break
        for j in range(2,num_rows):
            if ws.cell(row=j, column=k).value == "ISAM_384":
                if ws.cell(row=j,column=num_cols-1).value == ws.cell(row=j,column=num_cols).value :
                    continue
                else:
                    ws.cell(row=j, column=num_cols).fill = PatternFill(fill_type="solid", start_color=YELLOW)
                    ws.cell(row=j, column=num_cols-1).fill = PatternFill(fill_type="solid", start_color=YELLOW)
            else:
                ws.cell(row=j, column=num_cols).value = ""

        wb.save(filename1)



if __name__ == '__main__':
    t = test()
    t.f1()
    PG.alert(text='Updated CDA Loss Profile', title='Success')
