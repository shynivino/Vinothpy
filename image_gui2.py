
from kivy.config import Config
#from kivy.core.window import Window
#Config.set('graphics', 'maximize', 1)
Config.set('graphics', 'width', '1175')
Config.set('graphics', 'height', '720')
from kivy.app import App
from kivy.uix.label import Label
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.gridlayout import GridLayout
from kivy.uix.behaviors import ButtonBehavior
from kivy.uix.widget import Widget
from kivy.graphics import BorderImage
import  openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
from kivy.clock import Clock
#import win32gui
count = 1
import os
import webbrowser 
from kivy.graphics import *
import time
from kivy.app import App
from kivy.uix.button import Button
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.gridlayout import GridLayout
from kivy.uix.label import Label
from kivy.uix.textinput import TextInput


class ImageDisplay(BoxLayout):
    RowCount = 2
    wb = openpyxl.load_workbook("C:\\NBN\\tf_files\\2BTM-20\\TestOutput_FINAL.xlsx")
    sheet = wb.get_sheet_by_name('Sheet1')
    #print ("in image display")
    coordinates = sheet['F' + str(RowCount)].value.split("\\")[6].split(".jpg")[0]
    img_location = sheet['G' + str(RowCount)].value[1:]
    #print("the image loaction",img_location)
    img_name = sheet['E' + str(RowCount)].value
    auto_pole_type = sheet['H' + str(RowCount)].value
    auto_pole_score = sheet['I' + str(RowCount)].value
    complete = False
    
    #def __init__(self):
    #    self.stop = False
    def change_label1(self, *args):
        self.stop = False
        ##print ("calling change_label1")
        label2 = self.ids['label2']
        label3 = self.ids['co-ordinates']
        img = self.ids['main_image']
        img.source = "not_found.jpg"
        img.reload()
        self.get_total_rows()
        Clock.schedule_once(self.callback, 2)
    
         
       
        #Clock.schedule_interval(self.callback, 2)

    def get_total_rows(self):
        #print ("in get total")
        row_no = self.RowCount
        img_loc = self.sheet['G' + str(row_no)].value
        while img_loc != None:
            row_no += 1
            ##print row_no,img_loc
            img_loc = self.sheet['G' + str(row_no)].value
        self.total_row_count = row_no-1
        ##print "exit get total"

    def stop_image(self):
        self.stop = True

    def report(self):
        self.report = True
        url= "file:///C:\\NBN\\PoleCleansing\\tabs-frames\\tabs-frames-demo.html"
        webbrowser.open(url)

    def next_image(self):
        self.stop = False
        
        if self.complete == True:
            
            self.RowCount = 2
            self.img_location = self.sheet['G' + str(self.RowCount)].value[1:]
            self.auto_pole_type = self.sheet['H' + str(self.RowCount)].value
            self.auto_pole_score = self.sheet['I' + str(self.RowCount)].value
            Clock.schedule_interval(self.callback, 2)
        else:
            Clock.schedule_interval(self.callback, 2)

        Clock.schedule_once(self.callback, 2)

    def callback(self, dt):
        ##print ("calling callback")
        if self.stop == False and self.img_location != None:
            try:
                #print ("cccccccc", self.img_location)
                if self.RowCount != self.sheet.max_row:
                    #self.img_location = self.img_location[1:]
                    ##print("in callback",self.img_location)
                    if os.path.exists(self.img_location):
                        #print ("working")
                        img = self.ids['main_image']
                        img.source = self.img_location
                        img.reload()
                        label3 = self.ids['co-ordinates']
                        label3.text = self.coordinates
                        ##print(label3.text)
                        label_image1=self.ids['noneimage']
                        label_image1.source =".\\images\\1.jpg"
                        label_image2=self.ids['straight jointuse']
                        label_image2.source =".\\images\\2.jpg"
                        label_image3=self.ids['telstra only']
                        label_image3.source =".\\images\\3.jpg"
                        label_image4=self.ids['utilitybulb']
                        label_image4.source =".\\images\\4.jpg"
                        label_image5=self.ids['utilityonly']
                        label_image5.source =".\\images\\5.jpg"
                        label_image6=self.ids['t joint jointuse']
                        label_image6.source =".\\images\\6.jpg"
                        
                        
                        try:
                            if self.auto_pole_type=='noneimage':
                                label_image1.source =".\\images\\1_borderimage.jpg"
                                label_image2.source =".\\images\\2.jpg"
                                label_image3.source =".\\images\\3.jpg"
                                label_image4.source =".\\images\\4.jpg"
                                label_image5.source =".\\images\\5.jpg"
                                label_image6.source =".\\images\\6.jpg"
                                              
                            elif self.auto_pole_type=='straight jointuse':
                                label_image2.source =".\\images\\2_borderimage.jpg"
                                label_image1.source =".\\images\\1.jpg"
                                label_image3.source =".\\images\\3.jpg"
                                label_image6.source =".\\images\\6.jpg"
                                label_image5.source =".\\images\\5.jpg"
                                label_image4.source =".\\images\\4.jpg"
                                
                                
                            elif self.auto_pole_type=='telstra only':
                                label_image3.source =".\\images\\3_borderimage.jpg"
                                label_image1.source =".\\images\\1.jpg"
                                label_image2.source =".\\images\\2.jpg"
                                label_image5.source =".\\images\\5.jpg"
                                label_image4.source =".\\images\\4.jpg"
                                label_image6.source =".\\images\\6.jpg"

                            elif self.auto_pole_type=='utilitybulb':
                                label_image4.source =".\\images\\4_borderimage.jpg"
                                label_image1.source =".\\images\\1.jpg"
                                label_image3.source =".\\images\\3.jpg"
                                label_image5.source =".\\images\\5.jpg"
                                label_image2.source =".\\images\\2.jpg"
                                label_image6.source =".\\images\\6.jpg"
                                
                                
                            elif self.auto_pole_type=='utilityonly':
                                label_image5.source =".\\images\\5_borderimage.jpg"
                                label_image1.source =".\\images\\1.jpg"
                                label_image2.source =".\\images\\2.jpg"
                                label_image4.source =".\\images\\4.jpg"
                                label_image3.source =".\\images\\3.jpg"
                                label_image6.source =".\\images\\6.jpg"
                                

                            elif self.auto_pole_type=='t joint jointuse':
                                label_image6.source =".\\images\\6_borderimage.jpg"
                                label_image1.source =".\\images\\1.jpg"
                                label_image3.source =".\\images\\3.jpg"
                                label_image2.source =".\\images\\2.jpg"
                                label_image4.source =".\\images\\4.jpg"
                                label_image5.source =".\\images\\5.jpg"
                                

                        except:
                            pass
                        lab1 = self.ids['label1']
                        if self.auto_pole_type == 'noneimage':
                            self.auto_pole_type = 'Nopole'
                            lab1.text = self.auto_pole_type
                            #print(lab1.text )
                        elif self.auto_pole_type == 'No Pole':
                            self.auto_pole_type = 'Nopole'
                            lab1.text = self.auto_pole_type
                            #print(lab1.text )
                        elif self.auto_pole_type == 'straight jointuse':
                             self.auto_pole_type = 'Broadband pole'
                             lab1.text = self.auto_pole_type
                        elif self.auto_pole_type == 't joint jointuse':
                             self.auto_pole_type = 't joint pole'
                             lab1.text = self.auto_pole_type
                        elif self.auto_pole_type == 'telstra only':
                             self.auto_pole_type = 'telstra pole'
                             lab1.text = self.auto_pole_type
                        #print ("0000000", self.auto_pole_score, type(self.auto_pole_score))
                        auto_pole_score_long = round(float(self.auto_pole_score) * 100,2)
    
                        self.auto_pole_score = str(auto_pole_score_long) 
                       
                        lab2 = self.ids['label2']
                        lab2.text = self.auto_pole_score
                        #print ("second occurence")
                        self.RowCount += 1
                        try:
                            self.coordinates = self.sheet['F' + str(self.RowCount)].value.split("\\")[6].split(".jpg")[0]
                            self.img_location = self.sheet['G' + str(self.RowCount)].value[1:]
                            self.auto_pole_type = self.sheet['H' + str(self.RowCount)].value
                            self.auto_pole_score = self.sheet['I' + str(self.RowCount)].value
                        except:
                            pass
                    else:
                        img = self.ids['main_image']
                        img.source = "not_found.jpg"
                        img.reload()
                        img_name = self.ids['image_not_found']
                        img_name.text = "Image Name " +self.sheet['E' + str(self.RowCount)].value + "\n not found."
                        lab1 = self.ids['label1']
                        #lab1.text = "Auto_Pole Type\n" + self.auto_pole_type
                        lab1.text =  self.auto_pole_type
                        lab2 = self.ids['label2']
                        #lab2.text = "Auto_Pole_Score\n" + self.auto_pole_score
                        lab2.text =  self.auto_pole_score
                        label3 = self.ids['co-ordinates']
                        label3.text = self.coordinates
                        self.RowCount += 1
                        #print ("third occurence")
                        self.coordinates = self.sheet['F' + str(self.RowCount)].value.split("\\")[6].split(".jpg")[0]
                        self.img_location = self.sheet['G' + str(self.RowCount)].value[1:]
                        self.auto_pole_type = self.sheet['H' + str(self.RowCount)].value
                        self.auto_pole_score = self.sheet['I' + str(self.RowCount)].value
                elif self.img_location == None:
                    ##print ("end")
                    img = self.ids['main_image']
                    img.source = "end.jpg"
                    img.reload()
                    #label3 = self.ids['co-ordinates']
                    #label3.text = self.coordinates
                        
                    

                else:
                    ##print ("in else")
                    #print (self.img_location)
                    #print ("+++++++++++++",self.img_location)
                    self.RowCount = 2
                    #print("Last occurence")
                    self.coordinates = self.sheet['F' + str(self.RowCount)].value.split("\\")[6].split(".jpg")[0]
                    self.img_location = self.sheet['G' + str(self.RowCount)].value[1:]
                    self.auto_pole_type = self.sheet['H' + str(self.RowCount)].value
                    self.auto_pole_score = self.sheet['I' + str(self.RowCount)].value
                    self.next_image()
                    return False

            except Exception as e:
                #print(str(e))
                img = self.ids['main_image']
                img.source = "not_found.jpg"
                img.reload()
        else:
            ##print "in else unscedule"
            Clock.unschedule(self.callback)
        progress_bar = self.ids['pb']
        ##print "row no",self.RowCount

##        print ("row" , ((self.RowCount-1)/float(self.total_row_count)) *100)
##        print("ror2",self.RowCount,"row",float(self.total_row_count))
        
        progress_bar.value = ((self.RowCount-3)/float(self.total_row_count)) *100
        '''
        #print ("----->",self.img_location)
        if self.RowCount == self.sheet.max_row:
            #print ("+++++++++++++",self.img_location)
            self.RowCount = 2
            #print("Last occurence")
            self.coordinates = self.sheet['F' + str(self.RowCount)].value.split("\\")[6].split(".jpg")[0]
            self.img_location = self.sheet['G' + str(self.RowCount)].value[1:]
            self.auto_pole_type = self.sheet['H' + str(self.RowCount)].value
            self.auto_pole_score = self.sheet['I' + str(self.RowCount)].value
            self.next_image()
        '''

        

class MainApp(App):
    def build(self):
        self.title ='AI Based Network Reconciliation'
        x = ImageDisplay()
        x.change_label1()
        return x
        y = ImageDisplay()
        return y
        
    '''
    def on_start(self):
        handle = win32gui.FindWindow(0, "main")
        win32gui.SetForegroundWindow(handle)
    '''
if __name__ == "__main__":
    MainApp().run()
   
